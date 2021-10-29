import pandas
import openpyxl
from openpyxl import Workbook
from pandas import DataFrame
# 打开工作簿
wb = openpyxl.load_workbook(r'C:\Users\27769\Desktop\python\term1.xlsx')
sheet = wb.active

row_num = sheet.max_row # 获取当前表中最大的行数
course_number=1
sum=0
all_gpa=0
p=1
name=[]
all_student_gpa=[]
for row in range(1, row_num+1):
    full_gpa = sheet.cell(row, 10).value #满绩点所在excel表格的列数
    gpa =sheet.cell(row,13).value #个人绩点所在excel表格的列数
    if full_gpa != None and gpa != None:
        if row == 1:

            sum += full_gpa * gpa
            all_gpa += full_gpa
        elif sheet.cell(row, 2).value == sheet.cell(row - 1, 2).value:
                sum += full_gpa * gpa
                all_gpa += full_gpa

        else:
            ave_gpa = sum / all_gpa
            print(sheet.cell(row - 1, 2).value + "的均绩为:", ave_gpa)
            name.append(sheet.cell(row - 1, 2).value)
            all_student_gpa.append(ave_gpa)
            sum = 0
            all_gpa = 0
            sum += full_gpa * gpa
            all_gpa += full_gpa

data={'姓名':name,'均绩':all_student_gpa}
df=DataFrame(data)
df.to_excel('result.xlsx')
#ok


